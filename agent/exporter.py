"""
Excel export.

Responsibilities:
- Write extracted call data to an Excel file.
- Schema: url, objective, inclusion_criteria, exclusion_criteria, deadline,
  max_funding, max_duration, procedure, contact, misc, remarks.
"""
import re
import math
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from typing import Any


COLUMNS = [
    'url',
    'objective',
    'inclusion_criteria',
    'exclusion_criteria',
    'deadline',
    'max_funding',
    'max_duration',
    'procedure',
    'contact',
    'misc',
    'remarks',
]

_FILL_GREEN   = PatternFill(fill_type='solid', fgColor='C6EFCE')
_FILL_RED     = PatternFill(fill_type='solid', fgColor='FFC7CE')
_ALIGN_CENTER = Alignment(wrap_text=True, vertical='center', horizontal='center')

# Points per line of text at the default font size (approx. 11pt Calibri)
_LINE_HEIGHT_PT = 14.5
# Minimum row height in points
_MIN_ROW_HEIGHT = 15


def _contains_any(text: str, keywords: list[str]) -> bool:
    """
    Return True if any keyword appears as a whole word in text (case-insensitive).

    Word-boundary matching prevents short terms like 'KI' or 'AI' from matching
    inside longer words (e.g. 'making', 'sustainable').

    A trailing 's' is made optional so that singular keywords (e.g. 'research
    institution', 'Forschungseinrichtung') also match their plural forms
    ('research institutions', 'Forschungseinrichtungen' is handled separately,
    but simple English/German plurals ending in 's' are covered automatically).
    """
    for kw in keywords:
        pattern = r'\b' + re.escape(kw) + r's?\b'
        if re.search(pattern, text, re.IGNORECASE):
            return True
    return False


def _estimate_row_height(row_texts: list[str], col_widths: list[float]) -> float:
    """
    Estimate the row height in points needed to display all cell contents without
    clipping, given each cell's text and column width (in Excel character units).

    The approximation: lines_needed = ceil(len(text) / effective_chars_per_line),
    where effective_chars_per_line ≈ col_width * 0.9 to account for variable-width
    characters and padding.
    """
    max_lines = 1
    for text, width in zip(row_texts, col_widths):
        if not text:
            continue
        # Count explicit newlines and estimate wrapped lines per paragraph
        paragraphs = str(text).split('\n')
        lines = 0
        chars_per_line = max(width * 0.9, 1)
        for para in paragraphs:
            lines += max(math.ceil(len(para) / chars_per_line), 1)
        max_lines = max(max_lines, lines)
    return max(max_lines * _LINE_HEIGHT_PT + 4, _MIN_ROW_HEIGHT)


def export_to_excel(records: list[dict], output_path: str, config: dict[str, Any] | None = None) -> None:
    """
    Write a list of call records to an Excel file.

    Each record should contain the keys defined in COLUMNS. Missing keys are
    filled with an empty string. Extra keys are ignored.

    Conditional background colouring (configured via config['excel_formatting']):
      - objective           → light green if any objective_keywords match
      - inclusion_criteria  → light green if any inclusion_keywords match
      - exclusion_criteria  → light red   if any exclusion_keywords match
    """
    fmt = (config or {}).get('excel_formatting', {})
    objective_kws: list[str] = fmt.get('objective_keywords',  [])
    inclusion_kws: list[str] = fmt.get('inclusion_keywords',  [])
    exclusion_kws: list[str] = fmt.get('exclusion_keywords',  [])

    rows = [{col: record.get(col, '') for col in COLUMNS} for record in records]
    df = pd.DataFrame(rows, columns=COLUMNS)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Bekanntmachungen')

        ws = writer.sheets['Bekanntmachungen']

        # --- Column widths: fit to widest cell, capped at 60 ---
        col_widths: dict[int, float] = {}
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            width = min(max_len + 2, 60)
            ws.column_dimensions[col_cells[0].column_letter].width = width
            col_widths[col_cells[0].column] = width

        # --- Header row: bold + centered ---
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = _ALIGN_CENTER

        # --- Data rows: centered, content-fitted height, conditional fills ---
        for row in ws.iter_rows(min_row=2):
            # Collect texts and widths for this row to estimate height
            row_texts  = [str(cell.value) if cell.value is not None else '' for cell in row]
            row_widths = [col_widths.get(cell.column, 10) for cell in row]
            ws.row_dimensions[row[0].row].height = _estimate_row_height(row_texts, row_widths)

            for cell in row:
                cell.alignment = _ALIGN_CENTER

                col_name = ws.cell(row=1, column=cell.column).value
                text = str(cell.value) if cell.value is not None else ''

                if col_name == 'objective' and objective_kws and _contains_any(text, objective_kws):
                    cell.fill = _FILL_GREEN
                elif col_name == 'inclusion_criteria' and inclusion_kws and _contains_any(text, inclusion_kws):
                    cell.fill = _FILL_GREEN
                elif col_name == 'exclusion_criteria' and exclusion_kws and _contains_any(text, exclusion_kws):
                    cell.fill = _FILL_RED
